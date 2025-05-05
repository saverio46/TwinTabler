import os
import io
import pandas as pd
import logging
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "default_secret_key_for_dev")
# Disable template caching during development
app.config['TEMPLATES_AUTO_RELOAD'] = True

# Define allowed file extensions
ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    """Check if the file has an allowed extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel(file_stream):
    """
    Process the Excel file according to requirements:
    1. Split data into odd-ID and even-ID rows
    2. Keep headers intact for odd-ID rows
    3. Append "_2" to headers for even-ID rows
    4. Position odd-ID rows starting from Column A
    5. Position even-ID rows starting from Column G
    """
    try:
        # Read Excel file
        df = pd.read_excel(file_stream)
        
        # Check if data exists
        if df.empty:
            return None, "The uploaded Excel file is empty"
        
        # Check if first column is numeric ID
        first_col_name = df.columns[0]
        if not pd.api.types.is_numeric_dtype(df[first_col_name]):
            return None, f"The first column '{first_col_name}' does not contain numeric IDs"
        
        # Split data into odd and even ID rows
        odd_rows = df[df[first_col_name] % 2 == 1]
        even_rows = df[df[first_col_name] % 2 == 0]
        
        # Rename headers for even rows
        even_rows.columns = [f"{col}_2" for col in even_rows.columns]
        
        # Create a new Excel workbook
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        # Create a blank DataFrame with combined columns
        combined_columns = list(odd_rows.columns) + list(even_rows.columns)
        combined_df = pd.DataFrame(columns=combined_columns)
        
        # Write to Excel with specified column positions
        combined_df.to_excel(writer, index=False, sheet_name='Processed Data')
        
        # Access the workbook
        workbook = writer.book
        worksheet = workbook['Processed Data']
        
        # Remove the empty header row that was created
        worksheet.delete_rows(1)
        
        # Write the headers
        for col_idx, col_name in enumerate(odd_rows.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = col_name
            
        for col_idx, col_name in enumerate(even_rows.columns, start=len(odd_rows.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = col_name
        
        # Write the data for odd rows
        for row_idx, row in enumerate(dataframe_to_rows(odd_rows, index=False, header=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
        
        # Write the data for even rows (starting from column G or next available column)
        start_col = len(odd_rows.columns) + 1
        for row_idx, row in enumerate(dataframe_to_rows(even_rows, index=False, header=False), start=2):
            for col_idx, value in enumerate(row, start=start_col):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = value
        
        # Save the Excel file
        workbook.save(output)
        output.seek(0)
        
        return output, None
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return None, f"Error processing Excel file: {str(e)}"

@app.route('/', methods=['GET', 'POST'])
def index():
    """Handle the main page and file uploads"""
    if request.method == 'POST':
        # Check if file part exists in the request
        if 'file' not in request.files:
            flash('No file part in the request', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        
        # Check if user selected a file
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        # Check if file has allowed extension
        if file and allowed_file(file.filename):
            try:
                # Secure the filename
                filename = secure_filename(file.filename)
                
                # Process the Excel file
                processed_file, error = process_excel(file)
                
                if error:
                    flash(error, 'danger')
                    return redirect(request.url)
                
                # Generate output filename
                output_filename = f"processed_{filename}"
                
                # Return the processed file for download
                return send_file(
                    processed_file,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
            except Exception as e:
                logger.error(f"Error during file processing: {str(e)}")
                flash(f"An error occurred during file processing: {str(e)}", 'danger')
                return redirect(request.url)
        else:
            flash('Only .xlsx files are allowed', 'danger')
            return redirect(request.url)
    
    # GET request - display the upload form
    return render_template('index.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
